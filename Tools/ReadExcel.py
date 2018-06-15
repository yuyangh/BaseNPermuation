# -*- coding: utf-8 -*-

# read and write 2007 excel
import openpyxl
import json
import os


# filePath="Data/正确.xlsx"

def read07Excel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb['Sheet1']
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


'''
@:return a list of map {name:string,url:string,JSONdata:json} from Excel document
'''
def getAllNameJSONDataURL(path):
    data = []
    wb = openpyxl.load_workbook(path)
    sheet = wb['Sheet1']
    for row in sheet.rows:
        if (len(row)) >= 8:
            name = row[2].value
            url = row[7].value
            JSONdata = json.loads(row[8].value)
            data.append({"name": name, 'url': url, 'JSONdata': JSONdata})
    return data


'''
@:parameter path: excel document path 
@:return a list of map {name:string, url:string, JSONdata:json}'''
def getNameURLList(path):
    data = []
    wb = openpyxl.load_workbook(path)
    sheet = wb['图片识别情况']
    for row in sheet.rows:
        if (len(row)) >= 3:
            name = row[2].value
            url = row[0].value
            JSONdata = ""
            data.append({"name": name, 'url': url, 'JSONdata': JSONdata})
    return data

